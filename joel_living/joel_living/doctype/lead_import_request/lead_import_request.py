# # Copyright (c) 2025, Subburaj and contributors
# # For license information, please see license.txt

# import frappe
# from frappe.utils import get_url, format_datetime
# from frappe.model.document import Document
# from frappe.utils.background_jobs import enqueue
# from joel_living.email import send_custom_email
# import pandas as pd
# import os
# from datetime import datetime


# class LeadImportRequest(Document):
#     def after_insert(self):
#         """Enqueue email notification and system notification after lead import request is created."""
#         enqueue(self.send_notifications, queue='short', job_name=f"lead_import_notify_{self.name}")

#     def send_notifications(self):
#         """Send both email and system notification."""
#         admin_settings = frappe.get_single("Admin Settings")

#         manager_roles = ["Admin", "Super Admin"]
#         recipients_list = frappe.db.sql("""
#             SELECT DISTINCT hr.parent
#             FROM `tabHas Role` AS hr
#             INNER JOIN `tabUser` AS u ON hr.parent = u.name
#             WHERE
#                 hr.role IN %(roles)s
#                 AND u.enabled = 1
#         """, {"roles": manager_roles}, as_list=1)
        
#         recipients = [row[0] for row in recipients_list]

#         if not recipients:
#             frappe.log_error("No active 'Admin' or 'Super Admin' users found for notification.", "Lead Import Notification Skipped")
#             return

#         requester = frappe.utils.get_fullname(self.owner) or self.owner
#         doc_url = f"{get_url()}/app/lead-import-request/{self.name}"
#         subject = f"New Lead Import Request from {requester}"
#         should_send_email = frappe.db.get_single_value("Admin Settings", "send_mail_on_import_and_export_request")
        
#         # ----- Send Email -----
#         if admin_settings.email_account and should_send_email:
#             # cc_emails = []
#             # if admin_settings.cc_mail_ids:
#             #     cc_emails = [e.strip() for e in admin_settings.cc_mail_ids.replace("\n", ",").split(",") if e.strip()]

#             message = f"""
#             <div style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #f7f9fc; padding: 25px;">
#                 <div style="max-width: 600px; margin: auto; background-color: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); padding: 25px;">
#                     <h2 style="color: #2c3e50; text-align:center; border-bottom: 2px solid #3498db; padding-bottom: 10px;">
#                         🆕 New Lead Import Request
#                     </h2>

#                     <p>A new <strong>Lead Import Request</strong> has been raised by 
#                         <b>{requester}</b>.
#                     </p>

#                     <div style="text-align:center; margin: 25px 0;">
#                         <a href="{doc_url}" style="
#                             background-color: #3498db;
#                             color: white;
#                             text-decoration: none;
#                             padding: 12px 22px;
#                             border-radius: 5px;
#                             font-weight: 500;
#                             font-size: 15px;
#                             display: inline-block;">
#                             🔍 View Lead Import Request
#                         </a>
#                     </div>

#                     <div style="background-color: #f9fafc; padding: 12px 18px; border-left: 3px solid #3498db; font-size: 14px; color: #555;">
#                         <p><strong>Request ID:</strong> {self.name}</p>
#                         <p><strong>Status:</strong> {self.status or "Draft"}</p>
#                         <p><strong>Submitted By:</strong> {requester}</p>
#                         <p><strong>Submitted On:</strong> {format_datetime(self.creation, "dd-MMM-yyyy HH:mm")}</p> 
#                     </div>

#                     <p style="margin-top:25px; color:#95a5a6; font-size:13px; text-align:center;">
#                         This is an automated notification from your system.
#                     </p>
#                 </div>
#             </div>
#             """

#             send_custom_email(
#                 to=recipients,
#                 # cc=cc_emails,
#                 subject=subject,
#                 message=message
#             )
#         else:
#             frappe.log_error("⚠️ Admin Settings missing Email Account or To Mail ID.", "Lead Import Notification Skipped")

#         # ----- Send System Notification -----
#         for user in recipients:
#             try:
#                 frappe.publish_realtime(
#                     event="msgprint",
#                     message=f"🆕 New Lead Import Request: {self.name} submitted by {requester}.",
#                     user=user
#                 )

#                 # Only create Notification Log if user exists
#                 # if frappe.get_value("User", admin_settings.to_mail_id):
#                 frappe.get_doc({
#                     "doctype": "Notification Log",
#                     "document_type": self.doctype,
#                     "document_name": self.name,
#                     "subject": subject,
#                     "from_user": self.owner,
#                     "for_user": user,
#                     "type": "Alert"
#                 }).insert(ignore_permissions=True)

#             except Exception as e:
#                 # Log the error but do not block the job
#                 frappe.log_error(e, "Lead Import Request Notification Skipped")






# @frappe.whitelist()
# def approve_request(docname):
#     # Update status to Importing...
#     doc = frappe.get_doc("Lead Import Request", docname)
#     doc.status = "Importing"
#     doc.save(ignore_permissions=True)
#     frappe.db.commit()

#     # Enqueue background job
#     frappe.enqueue(
#         "joel_living.joel_living.doctype.lead_import_request.lead_import_request.process_lead_import",
#         queue="long",   # for large files
#         timeout=3600,   # adjust for very large Excel
#         docname=docname,
#     )

#     return {"status": "queued", "message": "Lead import started in background."}


# def process_lead_import(docname):
#     doc = frappe.get_doc("Lead Import Request", docname)
#     print("Processing lead import for:", doc.owner)
#     added_count = 0
#     failed_count = 0
#     failed_rows = []
#     failed_reasons_set = set()
#     owner_email = doc.owner

#     try:
#         # Get file path
#         file_doc = frappe.get_doc("File", {"file_url": doc.file})
#         filepath = file_doc.get_full_path()

#         # Read Excel
#         df = pd.read_excel(filepath)

#         # Import leads
#         for _, row in df.iterrows():
#             try:
#                 lead = frappe.get_doc({
#                     "doctype": "Lead",
#                     "first_name": row.get("First Name"),
#                     "middle_name": row.get("Middle Name"),
#                     "last_name": row.get("Last Name"),
#                     "custom_genders": row.get("Gender"),
#                     "custom_budget_range":row.get("Budget Range"),
#                     "custom_budget_usd": row.get("Budget Value (USD)"),
#                     "custom_priority_": row.get("Priority"),
#                     "custom_project": row.get("Project"),
#                     "custom_developer": row.get("Developer"),
#                     "custom_developer_representative": row.get("Developer Representative"),
#                     "custom_lead_stages": row.get("Lead Stage"),
#                     "custom_main_lead_source": row.get("Main Lead Source"),
#                     "custom_secondary_lead_sources": row.get("Secondary Lead Source"),
#                     "custom_lead_type": row.get("Lead Type"),
#                     "job_title": row.get("Job Title"),
#                     "mobile_no": row.get("Mobile No"),
#                     "phone_ext":row.get("Secondary Phone"),
#                     "whatsapp_no": row.get("WhatsApp"),
#                     "email_id": row.get("Email"),
#                     "country": row.get("Country"),
#                     "state": row.get("State/Province"),
#                     "city": row.get("City"),
#                     "custom_lead_category":"Fresh Lead",
#                     "lead_owner": doc.owner,
#                     "custom_assigned_at":datetime.now(),
#                 })
#                 lead.insert(ignore_permissions=True)
#                 added_count += 1
#             except Exception as e:
#                 failed_count += 1
#                 failed_rows.append(row)
#                 failed_reasons_set.add(str(e))

#         # Save counts and reasons
#         doc.success_count = added_count
#         doc.failed_count = failed_count
#         doc.failed_reasons = ", ".join(failed_reasons_set)

#         # Generate Excel for failed leads if any
#         if failed_rows:
#             failed_df = pd.DataFrame(failed_rows)
#             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#             failed_filepath = os.path.join(
#                 frappe.get_site_path("private", "files"),
#                 f"failed_leads_{docname}_{timestamp}.xlsx"
#             )
#             failed_df.to_excel(failed_filepath, index=False)

#             # Attach file
#             file_doc = frappe.get_doc({
#                 "doctype": "File",
#                 "file_name": f"failed_leads_{docname}_{timestamp}.xlsx",
#                 "attached_to_doctype": "Lead Import Request",
#                 "attached_to_name": docname,
#                 "file_url": f"/private/files/failed_leads_{docname}_{timestamp}.xlsx",
#                 "is_private": 0
#             })
#             file_doc.save(ignore_permissions=True)
#             doc.failed_leads_file = file_doc.file_url

#         # Update status
#         if failed_count == 0:
#             doc.status = "Approved"
#         elif added_count > 0:
#             doc.status = "Partially Approved"
#         else:
#             doc.status = "Failed"

#         doc.save(ignore_permissions=True)
#         frappe.db.commit()

#         frappe.publish_realtime(
#             "lead_import_progress",
#             {"docname": docname, "status": doc.status, "added": added_count, "failed": failed_count},
#             user=frappe.session.user
#         )

#         # Enqueue notification
#         enqueue(
#             "joel_living.joel_living.doctype.lead_import_request.lead_import_request.send_lead_import_notification",
#             queue="short",
#             timeout=300,
#             docname=docname,
#             status=doc.status,
#             added=added_count,
#             failed=failed_count
#         )


#     except Exception as e:
#         doc.status = "Failed"
#         doc.save(ignore_permissions=True)
#         frappe.db.commit()

#         frappe.publish_realtime(
#             "lead_import_progress",
#             {"docname": docname, "status": "Failed", "error": str(e)},
#             user=frappe.session.user
#         )
#         # Enqueue failure notification
#         enqueue(
#             "joel_living.joel_living.doctype.lead_import_request.lead_import_request.send_lead_import_notification",
#             queue="short",
#             timeout=300,
#             docname=docname,
#             status="Failed",
#             error=str(e)
#         )
    
# from openpyxl import load_workbook

# @frappe.whitelist()
# def preview_excel(docname):
#     doc = frappe.get_doc("Lead Import Request", docname)
#     try:
#         file_doc = frappe.get_doc("File", {"file_url": doc.file})
#         filepath = file_doc.get_full_path()

#         wb = load_workbook(filepath)
#         ws = wb.active

#         # Convert to HTML table
#         html = "<table border='1' style='border-collapse: collapse; width:100%'>"

#         for i, row in enumerate(ws.iter_rows(values_only=True)):
#             html += "<tr>"
#             for cell in row:
#                 tag = "th" if i == 0 else "td"
#                 html += f"<{tag} style='padding:4px'>{cell if cell else ''}</{tag}>"
#             html += "</tr>"

#         html += "</table>"
#         return html

#     except Exception as e:
#         return f"<p style='color:red'>Error reading file: {e}</p>"


# @frappe.whitelist()
# def preview_failed_leads(docname):
#     doc = frappe.get_doc("Lead Import Request", docname)
#     if not doc.failed_leads_file:
#         return "<p style='color:gray'>No failed leads</p>"

#     try:
#         # Get file path
#         file_doc = frappe.get_doc("File", {"file_url": doc.failed_leads_file})
#         filepath = file_doc.get_full_path()

#         wb = load_workbook(filepath)
#         ws = wb.active

#         # Convert to HTML table
#         html = "<table border='1' style='border-collapse: collapse; width:100%'>"
#         for i, row in enumerate(ws.iter_rows(values_only=True)):
#             html += "<tr>"
#             for cell in row:
#                 tag = "th" if i == 0 else "td"
#                 html += f"<{tag} style='padding:4px'>{cell if cell else ''}</{tag}>"
#             html += "</tr>"
#         html += "</table>"

#         return html
#     except Exception as e:
#         return f"<p style='color:red'>Error reading failed leads file: {e}</p>"
    
    
    
# @frappe.whitelist()
# def reject_request(docname, reason=None):
#     doc = frappe.get_doc("Lead Import Request", docname)
#     doc.status = "Rejected"
#     if reason:
#         doc.rejected_reason = reason
#     doc.save(ignore_permissions=True)
#     frappe.db.commit()
#     frappe.msgprint("Request Rejected")

#     # Send notifications
#     # Enqueue notification in background
#     enqueue(
#         "joel_living.joel_living.doctype.lead_import_request.lead_import_request.send_lead_import_notification",
#         queue="short",
#         timeout=300,
#         docname=docname,
#         status="Rejected"
#     )


# @frappe.whitelist()
# def move_to_pending(docname):
#     doc = frappe.get_doc("Lead Import Request", docname)
#     doc.status = "Pending"
#     doc.rejected_reason = ""  # clear previous reason if needed
#     doc.save(ignore_permissions=True)
#     frappe.db.commit()
#     frappe.msgprint("Lead Import Request is now Pending.")





# # -------------------------------
# # Common Notification Function
# # -------------------------------
# def send_lead_import_notification(docname, status, added=0, failed=0, error=None):
#     """Send email + system notification for Lead Import Request"""
#     try:
#         doc = frappe.get_doc("Lead Import Request", docname)
#         admin_settings = frappe.get_single("Admin Settings")

#         if not admin_settings.email_account or not doc.owner:
#             frappe.log_error(f"Skipping notification: missing email account or owner for {docname}", "Lead Import Notification")
#             return

#         owner_email = doc.owner
#         requester = frappe.utils.get_fullname(owner_email) or owner_email

#         # Validate owner email
#         if "@" not in owner_email:
#             frappe.log_error(f"Invalid owner email: {owner_email}", "Lead Import Notification Skipped")
#             return

#         # Prepare CC
#         cc_emails = []
#         if admin_settings.cc_mail_ids:
#             cc_emails = [e.strip() for e in admin_settings.cc_mail_ids.replace("\n", ",").split(",") if e.strip()]

#         doc_url = f"{get_url()}/app/lead-import-request/{doc.name}"
#         subject = f"Lead Import Request {status}: {doc.name}"

#         # Email message
#         message = f"""
#         <div style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #f7f9fc; padding: 25px;">
#             <div style="max-width: 600px; margin: auto; background-color: #fff; border-radius: 8px; 
#                         box-shadow: 0 2px 8px rgba(0,0,0,0.08); padding: 25px;">
#                 <h2 style="color: #2c3e50; text-align:center; border-bottom: 2px solid #3498db; padding-bottom: 10px;">
#                     Lead Import Request {status}
#                 </h2>
#                 <p style="font-size: 15px; color: #34495e;">
#                     Hello <b style="color:#2c3e50;">{requester}</b>, your Lead Import Request <b>{doc.name}</b> has <b>{status}</b>.
#                 </p>
#         """

#         if status in ["Approved", "Partially Approved"]:
#             message += f"""
#                 <div style="background-color: #f9fafc; padding: 15px; border-left: 4px solid #3498db; font-size: 14px; color: #555;">
#                     <p><strong>Status:</strong> {status}</p>
#                     <p><strong>Successfully Added:</strong> {added}</p>
#                     <p><strong>Failed:</strong> {failed}</p>
#                     <p><strong>Submitted On:</strong> {format_datetime(doc.creation, 'dd-MMM-yyyy HH:mm')}</p>
#                 </div>
#             """
#         elif status == "Rejected":
#             message += f"""
#                 <div style="background-color: #f9fafc; padding: 15px; border-left: 4px solid #e74c3c; font-size: 14px; color: #555;">
#                     <p><strong>Status:</strong> {status}</p>
#                     <p><strong>Reason:</strong> {doc.rejected_reason}</p>
#                     <p><strong>Submitted On:</strong> {format_datetime(doc.creation, 'dd-MMM-yyyy HH:mm')}</p>
#                 </div>
#             """
#         elif status == "Failed" and error:
#             message += f'<p style="color:#e74c3c;"><strong>Error:</strong> {error}</p>'

#         message += f"""
#                 <div style="text-align:center; margin: 25px 0;">
#                     <a href="{doc_url}" style="
#                         background-color: #3498db;
#                         color: white;
#                         text-decoration: none;
#                         padding: 12px 22px;
#                         border-radius: 5px;
#                         font-weight: 500;
#                         font-size: 15px;
#                         display: inline-block;">
#                         🔍 View Request
#                     </a>
#                 </div>
#                 <p style="margin-top:25px; color:#95a5a6; font-size:13px; text-align:center;">
#                     This is an automated notification from your system.
#                 </p>
#             </div>
#         </div>
#         """
#         should_send_email = frappe.db.get_single_value("Admin Settings", "send_mail_on_import_and_export_request")
#         if should_send_email:
#         # Send email
#             send_custom_email(to=owner_email, cc=cc_emails, subject=subject, message=message)

#         # Send system notification
#         description = f"Lead Import Request {doc.name} has been {status}."
#         if status == "Rejected":
#             description += f"\nReason: {doc.rejected_reason}"
#         if status == "Failed" and error:
#             description += f"\nError: {error}"

#         frappe.get_doc({
#             "doctype": "Notification Log",
#             "subject": f"Lead Import Request {status}: {doc.name}",
#             "type": "Alert",
#             "document_type": "Lead Import Request",
#             "document_name": doc.name,
#             "for_user": owner_email,
#             "description": description
#         }).insert(ignore_permissions=True)

#     except Exception as e:
#         frappe.log_error(f"Failed to send lead import notification: {str(e)}", "Lead Import Notification")











# Copyright (c) 2025, Subburaj and contributors
# For license information, please see license.txt
 
import frappe
from frappe.utils import get_url, format_datetime
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from joel_living.email import send_custom_email
import pandas as pd
import os
from datetime import datetime
 
 
class LeadImportRequest(Document):
    # def after_insert(self):
    #     """Enqueue email notification and system notification after lead import request is created."""
    #     enqueue(self.send_notifications, queue='short', job_name=f"lead_import_notify_{self.name}")
    def after_insert(self):
        """
        If uploader is Admin / Super Admin / Administrator → directly import.
        Else → send for approval.
        """
         # ✅ First validate Excel for everyone
        validate_excel_file(self)
 
        # ================================
        # ✅ NEW LOGIC : AUTO IMPORT FOR ADMINS
        # ================================
        admin_roles = ["Admin", "Super Admin", "Administrator"]
 
        user_roles = [r.role for r in frappe.get_all(
            "Has Role",
            filters={"parent": self.owner},
            fields=["role"]
        )]
 
        is_admin_user = any(role in admin_roles for role in user_roles)
       
        if is_admin_user:
            # ✅ CHANGED: Skip approval & start import directly
            self.status = "Importing"
            self.save(ignore_permissions=True)
            frappe.db.commit()
 
            frappe.enqueue(
                "joel_living.joel_living.doctype.lead_import_request.lead_import_request.process_lead_import",
                queue="long",
                timeout=3600,
                docname=self.name
            )
        else:
            # ✅ ORIGINAL BEHAVIOUR (non-admin users)
            enqueue(
                self.send_notifications,
                queue='short',
                job_name=f"lead_import_notify_{self.name}"
            )
 
    def send_notifications(self):
        """Send both email and system notification."""
        admin_settings = frappe.get_single("Admin Settings")
 
        manager_roles = ["Admin", "Super Admin"]
        recipients_list = frappe.db.sql("""
            SELECT DISTINCT hr.parent
            FROM `tabHas Role` AS hr
            INNER JOIN `tabUser` AS u ON hr.parent = u.name
            WHERE
                hr.role IN %(roles)s
                AND u.enabled = 1
        """, {"roles": manager_roles}, as_list=1)
       
        recipients = [row[0] for row in recipients_list]
 
        if not recipients:
            frappe.log_error("No active 'Admin' or 'Super Admin' users found for notification.", "Lead Import Notification Skipped")
            return
 
        requester = frappe.utils.get_fullname(self.owner) or self.owner
        doc_url = f"{get_url()}/app/lead-import-request/{self.name}"
        subject = f"New Lead Import Request from {requester}"
        should_send_email = frappe.db.get_single_value("Admin Settings", "send_mail_on_import_and_export_request")
       
        # ----- Send Email -----
        if admin_settings.email_account and should_send_email:
            # cc_emails = []
            # if admin_settings.cc_mail_ids:
            #     cc_emails = [e.strip() for e in admin_settings.cc_mail_ids.replace("\n", ",").split(",") if e.strip()]
 
            message = f"""
            <div style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #f7f9fc; padding: 25px;">
                <div style="max-width: 600px; margin: auto; background-color: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); padding: 25px;">
                    <h2 style="color: #2c3e50; text-align:center; border-bottom: 2px solid #3498db; padding-bottom: 10px;">
                        🆕 New Lead Import Request
                    </h2>
 
                    <p>A new <strong>Lead Import Request</strong> has been raised by
                        <b>{requester}</b>.
                    </p>
 
                    <div style="text-align:center; margin: 25px 0;">
                        <a href="{doc_url}" style="
                            background-color: #3498db;
                            color: white;
                            text-decoration: none;
                            padding: 12px 22px;
                            border-radius: 5px;
                            font-weight: 500;
                            font-size: 15px;
                            display: inline-block;">
                            🔍 View Lead Import Request
                        </a>
                    </div>
 
                    <div style="background-color: #f9fafc; padding: 12px 18px; border-left: 3px solid #3498db; font-size: 14px; color: #555;">
                        <p><strong>Request ID:</strong> {self.name}</p>
                        <p><strong>Status:</strong> {self.status or "Draft"}</p>
                        <p><strong>Submitted By:</strong> {requester}</p>
                        <p><strong>Submitted On:</strong> {format_datetime(self.creation, "dd-MMM-yyyy HH:mm")}</p>
                    </div>
 
                    <p style="margin-top:25px; color:#95a5a6; font-size:13px; text-align:center;">
                        This is an automated notification from your system.
                    </p>
                </div>
            </div>
            """
 
            send_custom_email(
                to=recipients,
                # cc=cc_emails,
                subject=subject,
                message=message
            )
        else:
            frappe.log_error("⚠️ Admin Settings missing Email Account or To Mail ID.", "Lead Import Notification Skipped")
 
        # ----- Send System Notification -----
        for user in recipients:
            try:
                frappe.publish_realtime(
                    event="msgprint",
                    message=f"🆕 New Lead Import Request: {self.name} submitted by {requester}.",
                    user=user
                )
 
                # Only create Notification Log if user exists
                # if frappe.get_value("User", admin_settings.to_mail_id):
                frappe.get_doc({
                    "doctype": "Notification Log",
                    "document_type": self.doctype,
                    "document_name": self.name,
                    "subject": subject,
                    "from_user": self.owner,
                    "for_user": user,
                    "type": "Alert"
                }).insert(ignore_permissions=True)
 
            except Exception as e:
                # Log the error but do not block the job
                frappe.log_error(e, "Lead Import Request Notification Skipped")
 
def validate_excel_file(doc):
    file_doc = frappe.get_doc("File", {"file_url": doc.file})
    filepath = file_doc.get_full_path()
 
    df = pd.read_excel(filepath)
 
    # No rows at all
    if df.empty:
        frappe.throw("Excel file is empty. Please upload a valid Excel file with lead data.")
 
    # Only headers but no data
    if len(df.index) == 0:
        frappe.throw("Excel file contains only headers and no lead data.")
 
import re
 
def is_valid_email(email):
    if not email:
        return False
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, str(email).strip())
 
def is_valid_mobile(mobile):
    if not mobile:
        return False
    mobile = str(mobile).strip()
    return mobile.isdigit() and 8 <= len(mobile) <= 15
 
 
 
 
@frappe.whitelist()
def approve_request(docname):
    # Update status to Importing...
    doc = frappe.get_doc("Lead Import Request", docname)
    doc.status = "Importing"
    doc.save(ignore_permissions=True)
    frappe.db.commit()
 
    # Enqueue background job
    frappe.enqueue(
        "joel_living.joel_living.doctype.lead_import_request.lead_import_request.process_lead_import",
        queue="long",   # for large files
        timeout=3600,   # adjust for very large Excel
        docname=docname,
    )
 
    return {"status": "queued", "message": "Lead import started in background."}
 
 
def process_lead_import(docname):
    doc = frappe.get_doc("Lead Import Request", docname)
    print("Processing lead import for:", doc.owner)
    added_count = 0
    failed_count = 0
    failed_rows = []
    failed_reasons_set = set()
    owner_email = doc.owner
 
    try:
        # Get file path
        file_doc = frappe.get_doc("File", {"file_url": doc.file})
        filepath = file_doc.get_full_path()
 
        # Read Excel
        df = pd.read_excel(filepath)
   
        # # Import leads
        # for _, row in df.iterrows():
        #     try:
        #         lead = frappe.get_doc({
        #             "doctype": "Lead",
        #             "first_name": row.get("First Name"),
        #             "middle_name": row.get("Middle Name"),
        #             "last_name": row.get("Last Name"),
        #             "custom_genders": row.get("Gender"),
        #             "custom_budget_range":row.get("Budget Range"),
        #             "custom_budget_usd": row.get("Budget Value (USD)"),
        #             "custom_priority_": row.get("Priority"),
        #             "custom_project": row.get("Project"),
        #             "custom_developer": row.get("Developer"),
        #             "custom_developer_representative": row.get("Developer Representative"),
        #             "custom_lead_stages": row.get("Lead Stage"),
        #             "custom_main_lead_source": row.get("Main Lead Source"),
        #             "custom_secondary_lead_sources": row.get("Secondary Lead Source"),
        #             "custom_lead_type": row.get("Lead Type"),
        #             "job_title": row.get("Job Title"),
        #             "mobile_no": row.get("Mobile No"),
        #             "phone_ext":row.get("Secondary Phone"),
        #             "whatsapp_no": row.get("WhatsApp"),
        #             "email_id": row.get("Email"),
        #             "country": row.get("Country"),
        #             "state": row.get("State/Province"),
        #             "city": row.get("City"),
        #             "custom_lead_category":"Fresh Lead",
        #             "lead_owner": doc.owner,
        #             "custom_assigned_at":datetime.now(),
        #         })
        #         lead.insert(ignore_permissions=True)
        #         added_count += 1
        #     except Exception as e:
        #         failed_count += 1
        #         failed_rows.append(row)
        #         failed_reasons_set.add(str(e))
        for _, row in df.iterrows():
            row_errors = []
 
            email = row.get("Email")
            mobile = row.get("Mobile No")
 
            # Validate Email
            if not is_valid_email(email):
                row_errors.append(f"Invalid Email: {email}")
 
            # Validate Mobile
            if not is_valid_mobile(mobile):
                row_errors.append(f"Invalid Mobile No: {mobile}")
 
            # If any validation failed → mark as failed
            if row_errors:
                failed_count += 1
                failed_rows.append(row)
                failed_reasons_set.update(row_errors)
                continue
 
            try:
                lead = frappe.get_doc({
                    "doctype": "Lead",
                    "first_name": row.get("First Name"),
                    "middle_name": row.get("Middle Name"),
                    "last_name": row.get("Last Name"),
                    "custom_genders": row.get("Gender"),
                    "custom_budget_range": row.get("Budget Range"),
                    "custom_budget_usd": row.get("Budget Value (USD)"),
                    "custom_priority_": row.get("Priority"),
                    "custom_project": row.get("Project"),
                    "custom_developer": row.get("Developer"),
                    "custom_developer_representative": row.get("Developer Representative"),
                    "custom_lead_stages": row.get("Lead Stage"),
                    "custom_main_lead_source": row.get("Main Lead Source"),
                    "custom_secondary_lead_sources": row.get("Secondary Lead Source"),
                    "custom_lead_type": row.get("Lead Type"),
                    "job_title": row.get("Job Title"),
                    "mobile_no": mobile,
                    "phone_ext": row.get("Secondary Phone"),
                    "whatsapp_no": row.get("WhatsApp"),
                    "email_id": email,
                    "country": row.get("Country"),
                    "state": row.get("State/Province"),
                    "city": row.get("City"),
                    "custom_lead_category": "Fresh Lead",
                    "lead_owner": doc.owner,
                    "custom_assigned_at": datetime.now(),
                })
 
                lead.insert(ignore_permissions=True)
                added_count += 1
 
            except Exception as e:
                failed_count += 1
                failed_rows.append(row)
                failed_reasons_set.add(str(e))
 
        # Save counts and reasons
        doc.success_count = added_count
        doc.failed_count = failed_count
        doc.failed_reasons = ", ".join(failed_reasons_set)
 
        # Generate Excel for failed leads if any
        if failed_rows:
            failed_df = pd.DataFrame(failed_rows)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            failed_filepath = os.path.join(
                frappe.get_site_path("private", "files"),
                f"failed_leads_{docname}_{timestamp}.xlsx"
            )
            failed_df.to_excel(failed_filepath, index=False)
 
            # Attach file
            file_doc = frappe.get_doc({
                "doctype": "File",
                "file_name": f"failed_leads_{docname}_{timestamp}.xlsx",
                "attached_to_doctype": "Lead Import Request",
                "attached_to_name": docname,
                "file_url": f"/private/files/failed_leads_{docname}_{timestamp}.xlsx",
                "is_private": 0
            })
            file_doc.save(ignore_permissions=True)
            doc.failed_leads_file = file_doc.file_url
 
        # Update status
        if failed_count == 0:
            doc.status = "Approved"
        elif added_count > 0:
            doc.status = "Partially Approved"
        else:
            doc.status = "Failed"
 
        doc.save(ignore_permissions=True)
        frappe.db.commit()
 
        frappe.publish_realtime(
            "lead_import_progress",
            {"docname": docname, "status": doc.status, "added": added_count, "failed": failed_count},
            user=frappe.session.user
        )
 
        # Enqueue notification
        enqueue(
            "joel_living.joel_living.doctype.lead_import_request.lead_import_request.send_lead_import_notification",
            queue="short",
            timeout=300,
            docname=docname,
            status=doc.status,
            added=added_count,
            failed=failed_count
        )
 
 
    except Exception as e:
        doc.status = "Failed"
        doc.save(ignore_permissions=True)
        frappe.db.commit()
 
        frappe.publish_realtime(
            "lead_import_progress",
            {"docname": docname, "status": "Failed", "error": str(e)},
            user=frappe.session.user
        )
        # Enqueue failure notification
        enqueue(
            "joel_living.joel_living.doctype.lead_import_request.lead_import_request.send_lead_import_notification",
            queue="short",
            timeout=300,
            docname=docname,
            status="Failed",
            error=str(e)
        )
   
from openpyxl import load_workbook
 
@frappe.whitelist()
def preview_excel(docname):
    doc = frappe.get_doc("Lead Import Request", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.file})
        filepath = file_doc.get_full_path()
 
        wb = load_workbook(filepath)
        ws = wb.active
 
        # Convert to HTML table
        html = "<table border='1' style='border-collapse: collapse; width:100%'>"
 
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            html += "<tr>"
            for cell in row:
                tag = "th" if i == 0 else "td"
                html += f"<{tag} style='padding:4px'>{cell if cell else ''}</{tag}>"
            html += "</tr>"
 
        html += "</table>"
        return html
 
    except Exception as e:
        return f"<p style='color:red'>Error reading file: {e}</p>"
 
 
@frappe.whitelist()
def preview_failed_leads(docname):
    doc = frappe.get_doc("Lead Import Request", docname)
    if not doc.failed_leads_file:
        return "<p style='color:gray'>No failed leads</p>"
 
    try:
        # Get file path
        file_doc = frappe.get_doc("File", {"file_url": doc.failed_leads_file})
        filepath = file_doc.get_full_path()
 
        wb = load_workbook(filepath)
        ws = wb.active
 
        # Convert to HTML table
        html = "<table border='1' style='border-collapse: collapse; width:100%'>"
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            html += "<tr>"
            for cell in row:
                tag = "th" if i == 0 else "td"
                html += f"<{tag} style='padding:4px'>{cell if cell else ''}</{tag}>"
            html += "</tr>"
        html += "</table>"
 
        return html
    except Exception as e:
        return f"<p style='color:red'>Error reading failed leads file: {e}</p>"
   
   
   
@frappe.whitelist()
def reject_request(docname, reason=None):
    doc = frappe.get_doc("Lead Import Request", docname)
    doc.status = "Rejected"
    if reason:
        doc.rejected_reason = reason
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    frappe.msgprint("Request Rejected")
 
    # Send notifications
    # Enqueue notification in background
    enqueue(
        "joel_living.joel_living.doctype.lead_import_request.lead_import_request.send_lead_import_notification",
        queue="short",
        timeout=300,
        docname=docname,
        status="Rejected"
    )
 
 
@frappe.whitelist()
def move_to_pending(docname):
    doc = frappe.get_doc("Lead Import Request", docname)
    doc.status = "Pending"
    doc.rejected_reason = ""  # clear previous reason if needed
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    frappe.msgprint("Lead Import Request is now Pending.")
 
 
 
 
 
# -------------------------------
# Common Notification Function
# -------------------------------
def send_lead_import_notification(docname, status, added=0, failed=0, error=None):
    """Send email + system notification for Lead Import Request"""
    try:
        doc = frappe.get_doc("Lead Import Request", docname)
        admin_settings = frappe.get_single("Admin Settings")
 
        if not admin_settings.email_account or not doc.owner:
            frappe.log_error(f"Skipping notification: missing email account or owner for {docname}", "Lead Import Notification")
            return
 
        owner_email = doc.owner
        requester = frappe.utils.get_fullname(owner_email) or owner_email
 
        # Validate owner email
        if "@" not in owner_email:
            frappe.log_error(f"Invalid owner email: {owner_email}", "Lead Import Notification Skipped")
            return
 
        # Prepare CC
        cc_emails = []
        if admin_settings.cc_mail_ids:
            cc_emails = [e.strip() for e in admin_settings.cc_mail_ids.replace("\n", ",").split(",") if e.strip()]
 
        doc_url = f"{get_url()}/app/lead-import-request/{doc.name}"
        subject = f"Lead Import Request {status}: {doc.name}"
 
        # Email message
        message = f"""
        <div style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #f7f9fc; padding: 25px;">
            <div style="max-width: 600px; margin: auto; background-color: #fff; border-radius: 8px;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.08); padding: 25px;">
                <h2 style="color: #2c3e50; text-align:center; border-bottom: 2px solid #3498db; padding-bottom: 10px;">
                    Lead Import Request {status}
                </h2>
                <p style="font-size: 15px; color: #34495e;">
                    Hello <b style="color:#2c3e50;">{requester}</b>, your Lead Import Request <b>{doc.name}</b> has <b>{status}</b>.
                </p>
        """
 
        if status in ["Approved", "Partially Approved"]:
            message += f"""
                <div style="background-color: #f9fafc; padding: 15px; border-left: 4px solid #3498db; font-size: 14px; color: #555;">
                    <p><strong>Status:</strong> {status}</p>
                    <p><strong>Successfully Added:</strong> {added}</p>
                    <p><strong>Failed:</strong> {failed}</p>
                    <p><strong>Submitted On:</strong> {format_datetime(doc.creation, 'dd-MMM-yyyy HH:mm')}</p>
                </div>
            """
        elif status == "Rejected":
            message += f"""
                <div style="background-color: #f9fafc; padding: 15px; border-left: 4px solid #e74c3c; font-size: 14px; color: #555;">
                    <p><strong>Status:</strong> {status}</p>
                    <p><strong>Reason:</strong> {doc.rejected_reason}</p>
                    <p><strong>Submitted On:</strong> {format_datetime(doc.creation, 'dd-MMM-yyyy HH:mm')}</p>
                </div>
            """
        elif status == "Failed" and error:
            message += f'<p style="color:#e74c3c;"><strong>Error:</strong> {error}</p>'
 
        message += f"""
                <div style="text-align:center; margin: 25px 0;">
                    <a href="{doc_url}" style="
                        background-color: #3498db;
                        color: white;
                        text-decoration: none;
                        padding: 12px 22px;
                        border-radius: 5px;
                        font-weight: 500;
                        font-size: 15px;
                        display: inline-block;">
                        🔍 View Request
                    </a>
                </div>
                <p style="margin-top:25px; color:#95a5a6; font-size:13px; text-align:center;">
                    This is an automated notification from your system.
                </p>
            </div>
        </div>
        """
        should_send_email = frappe.db.get_single_value("Admin Settings", "send_mail_on_import_and_export_request")
        if should_send_email:
        # Send email
            send_custom_email(to=owner_email, cc=cc_emails, subject=subject, message=message)
 
        # Send system notification
        description = f"Lead Import Request {doc.name} has been {status}."
        if status == "Rejected":
            description += f"\nReason: {doc.rejected_reason}"
        if status == "Failed" and error:
            description += f"\nError: {error}"
 
        frappe.get_doc({
            "doctype": "Notification Log",
            "subject": f"Lead Import Request {status}: {doc.name}",
            "type": "Alert",
            "document_type": "Lead Import Request",
            "document_name": doc.name,
            "for_user": owner_email,
            "description": description
        }).insert(ignore_permissions=True)
 
    except Exception as e:
        frappe.log_error(f"Failed to send lead import notification: {str(e)}", "Lead Import Notification")
 